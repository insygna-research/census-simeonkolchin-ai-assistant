"""
File content extraction for various text, document and spreadsheet formats.
"""

import io
import csv
import json
import logging
from typing import Dict, Any
from pathlib import Path

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".xml", ".html", ".htm",
    ".py", ".js", ".ts", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".log",
    ".docx", ".doc", ".pdf", ".xlsx",
}


class FileProcessor:
    """Extract text from uploaded files."""

    def process(self, filename: str, content: bytes) -> Dict[str, Any]:
        ext = Path(filename).suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            return {"error": f"Unsupported file type: {ext}"}

        try:
            if ext in (".txt", ".md", ".py", ".js", ".ts", ".yaml", ".yml",
                        ".toml", ".ini", ".cfg", ".log", ".xml", ".html", ".htm"):
                text = self._read_text(content)
            elif ext == ".csv":
                text = self._read_csv(content)
            elif ext == ".json":
                text = self._read_json(content)
            elif ext == ".docx":
                text = self._read_docx(content)
            elif ext == ".doc":
                text = self._read_doc(content)
            elif ext == ".pdf":
                text = self._read_pdf(content)
            elif ext == ".xlsx":
                text = self._read_xlsx(content)
            else:
                text = self._read_text(content)

            return {
                "success": True,
                "filename": filename,
                "extension": ext,
                "characters": len(text),
                "text": text,
            }
        except Exception as e:
            logger.error(f"Failed to process {filename}: {e}", exc_info=True)
            return {"error": f"Failed to process {filename}: {str(e)}"}

    def _read_text(self, content: bytes) -> str:
        for enc in ("utf-8", "cp1251", "latin-1"):
            try:
                return content.decode(enc)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="replace")

    def _read_csv(self, content: bytes) -> str:
        text = self._read_text(content)
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return text
        lines = []
        for row in rows[:500]:
            lines.append(" | ".join(row))
        result = "\n".join(lines)
        if len(rows) > 500:
            result += f"\n... (ещё {len(rows) - 500} строк)"
        return result

    def _read_json(self, content: bytes) -> str:
        text = self._read_text(content)
        try:
            data = json.loads(text)
            return json.dumps(data, ensure_ascii=False, indent=2)[:50000]
        except json.JSONDecodeError:
            return text

    def _read_docx(self, content: bytes) -> str:
        from docx import Document
        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def _read_doc(self, content: bytes) -> str:
        text_parts = []
        try:
            decoded = content.decode("utf-8", errors="ignore")
            for line in decoded.split("\n"):
                clean = line.strip()
                if clean and len(clean) > 2:
                    printable = "".join(c for c in clean if c.isprintable() or c in "\n\t")
                    if printable.strip():
                        text_parts.append(printable.strip())
        except Exception:
            pass
        return "\n".join(text_parts) if text_parts else "[Could not extract text from .doc file]"

    def _read_pdf(self, content: bytes) -> str:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=content, filetype="pdf")
        pages = []
        for page in doc:
            text = page.get_text()
            if text.strip():
                pages.append(text.strip())
        doc.close()
        return "\n\n---\n\n".join(pages)

    def _read_xlsx(self, content: bytes) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=500, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(cells))
            header = f"=== Лист: {sheet_name} ==="
            parts.append(header + "\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(parts)
